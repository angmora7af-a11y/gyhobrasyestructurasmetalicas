from __future__ import annotations

import io
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_audit
from app.core.database import get_db
from app.deps import get_current_user, require_role
from app.models.auth import AppUser
from app.models.catalog import Product
from app.models.import_batch import ImportBatch, ImportBatchRow
from app.schemas.catalog import (
    ImportBatchResponse,
    ImportRowError,
    ProductCreate,
    ProductResponse,
    ProductUpdate,
)

router = APIRouter(prefix="/products", tags=["catalog"])

REQUIRED_COLUMNS = ["sku_code", "description"]


@router.get("", response_model=list[ProductResponse])
async def list_products(
    active_only: bool = Query(True),
    db: AsyncSession = Depends(get_db),
    _user: AppUser = Depends(get_current_user),
):
    stmt = select(Product)
    if active_only:
        stmt = stmt.where(Product.is_active.is_(True))
    result = await db.execute(stmt.order_by(Product.sku_code))
    return result.scalars().all()


@router.post("", response_model=ProductResponse, status_code=status.HTTP_201_CREATED)
async def create_product(
    body: ProductCreate,
    db: AsyncSession = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "catalog_contributor")),
):
    product = Product(
        tenant_id=user.tenant_id,
        sku_code=body.sku_code,
        description=body.description,
        category=body.category,
        rental_enabled=body.rental_enabled,
        sale_enabled=body.sale_enabled,
        rental_price_per_day=body.rental_price_per_day,
        sale_price=body.sale_price,
        usage_notes=body.usage_notes,
    )
    db.add(product)
    await db.flush()
    await log_audit(
        db,
        tenant_id=user.tenant_id,
        user_id=user.id,
        entity_type="product",
        entity_id=product.id,
        action="create",
    )
    await db.commit()
    await db.refresh(product)
    return product


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user: AppUser = Depends(get_current_user),
):
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Product not found")
    return product


@router.patch("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: UUID,
    body: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "catalog_contributor")),
):
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Product not found")

    updates = body.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(product, field, value)

    await log_audit(
        db,
        tenant_id=user.tenant_id,
        user_id=user.id,
        entity_type="product",
        entity_id=product.id,
        action="update",
        payload=updates,
    )
    await db.commit()
    await db.refresh(product)
    return product


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_product(
    product_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
):
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Product not found")
    product.is_active = False
    await log_audit(
        db,
        tenant_id=user.tenant_id,
        user_id=user.id,
        entity_type="product",
        entity_id=product.id,
        action="soft_delete",
    )
    await db.commit()


@router.post("/import", response_model=ImportBatchResponse, status_code=status.HTTP_201_CREATED)
async def import_products(
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
    user: AppUser = Depends(require_role("admin", "catalog_contributor")),
):
    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Empty workbook")

    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c else "" for c in next(rows_iter, [])]

    for col in REQUIRED_COLUMNS:
        if col not in header:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Missing required column: {col}",
            )

    batch = ImportBatch(
        tenant_id=user.tenant_id,
        uploaded_by=user.id,
        file_name=file.filename or "upload.xlsx",
        status="draft",
        total_rows=0,
        error_rows=0,
    )
    db.add(batch)
    await db.flush()

    total = 0
    errors = 0
    for row_num, row_values in enumerate(rows_iter, start=2):
        total += 1
        row_data = dict(zip(header, row_values, strict=False))
        error_msg: str | None = None

        if not row_data.get("sku_code"):
            error_msg = "sku_code is required"
        elif not row_data.get("description"):
            error_msg = "description is required"

        if error_msg:
            errors += 1

        db.add(
            ImportBatchRow(
                batch_id=batch.id,
                row_number=row_num,
                raw_json=row_data,
                error_message=error_msg,
            )
        )

    batch.total_rows = total
    batch.error_rows = errors
    await db.commit()
    await db.refresh(batch)
    return batch


import_router = APIRouter(prefix="/import-batches", tags=["catalog"])


@import_router.get("/{batch_id}", response_model=ImportBatchResponse)
async def get_import_batch(
    batch_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user: AppUser = Depends(get_current_user),
):
    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Batch not found")
    return batch


@import_router.get("/{batch_id}/errors", response_model=list[ImportRowError])
async def get_import_batch_errors(
    batch_id: UUID,
    db: AsyncSession = Depends(get_db),
    _user: AppUser = Depends(get_current_user),
):
    result = await db.execute(
        select(ImportBatchRow)
        .where(
            ImportBatchRow.batch_id == batch_id,
            ImportBatchRow.error_message.isnot(None),
        )
        .order_by(ImportBatchRow.row_number)
    )
    rows = result.scalars().all()
    return [
        ImportRowError(row_number=r.row_number, error_message=r.error_message)
        for r in rows
    ]


@import_router.post("/{batch_id}/approve", response_model=ImportBatchResponse)
async def approve_import_batch(
    batch_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
):
    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Batch not found")
    if batch.status != "draft":
        raise HTTPException(status.HTTP_409_CONFLICT, f"Batch is already {batch.status}")

    result = await db.execute(
        select(ImportBatchRow).where(
            ImportBatchRow.batch_id == batch_id,
            ImportBatchRow.error_message.is_(None),
        )
    )
    valid_rows = result.scalars().all()

    for row in valid_rows:
        data = row.raw_json
        product = Product(
            tenant_id=user.tenant_id,
            sku_code=str(data.get("sku_code", "")),
            description=str(data.get("description", "")),
            category=data.get("category"),
            rental_enabled=bool(data.get("rental_enabled", False)),
            sale_enabled=bool(data.get("sale_enabled", False)),
            rental_price_per_day=data.get("rental_price_per_day"),
            sale_price=data.get("sale_price"),
            usage_notes=data.get("usage_notes"),
        )
        db.add(product)

    batch.status = "approved"
    await log_audit(
        db,
        tenant_id=user.tenant_id,
        user_id=user.id,
        entity_type="import_batch",
        entity_id=batch.id,
        action="approve",
        payload={"products_created": len(valid_rows)},
    )
    await db.commit()
    await db.refresh(batch)
    return batch


@import_router.post("/{batch_id}/reject", response_model=ImportBatchResponse)
async def reject_import_batch(
    batch_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: AppUser = Depends(require_role("admin")),
):
    batch = await db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Batch not found")
    if batch.status != "draft":
        raise HTTPException(status.HTTP_409_CONFLICT, f"Batch is already {batch.status}")

    batch.status = "rejected"
    await log_audit(
        db,
        tenant_id=user.tenant_id,
        user_id=user.id,
        entity_type="import_batch",
        entity_id=batch.id,
        action="reject",
    )
    await db.commit()
    await db.refresh(batch)
    return batch
