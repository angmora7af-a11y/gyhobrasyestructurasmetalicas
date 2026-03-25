#!/usr/bin/env python3
"""Genera DiccionarioDatos_GYH.xlsx desde el modelo lógico (database-model.dbml)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "DiccionarioDatos_GYH.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def add_sheet_headers(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    ws.freeze_panes = "A2"


def autosize(ws, max_width=55):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for row in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[letter].width = min(max(12, maxlen + 2), max_width)


def main():
    wb = Workbook()

    # --- Hoja 1: Diccionario de campos ---
    ws1 = wb.active
    ws1.title = "Diccionario_Campos"
    h1 = [
        "Tabla",
        "Columna",
        "Tipo_dato",
        "Restriccion",
        "Descripcion",
        "Referencia_FR_PDR",
    ]
    add_sheet_headers(ws1, h1)

    rows = [
        # tenant
        ("tenant", "id", "UUID", "PK, default gen_random_uuid()", "Identificador único de instancia/tenant.", "FR-110"),
        ("tenant", "code", "VARCHAR(32)", "UK, NOT NULL", "Código corto de la instancia (ej. BOG_MED).", "FR-110"),
        ("tenant", "name", "VARCHAR(255)", "NOT NULL", "Nombre legible de la instancia.", "FR-110"),
        ("tenant", "city_code", "VARCHAR(32)", "NULL", "Código ciudad para despliegues multi-ciudad.", "FR-110"),
        ("tenant", "timezone", "VARCHAR(64)", "Default America/Bogota", "Zona horaria para timestamps de negocio.", "—"),
        ("tenant", "settings_json", "JSONB", "NULL", "Branding, límites, flags (config por tenant).", "FR-110"),
        ("tenant", "created_at", "TIMESTAMPTZ", "Default now()", "Auditoría de creación.", "—"),
        # app_user
        ("app_user", "id", "UUID", "PK", "Usuario de aplicación.", "FR-001"),
        ("app_user", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Aislamiento multi-instancia.", "FR-001"),
        ("app_user", "email", "CITEXT", "NOT NULL, único por tenant", "Correo login (insensible a mayúsculas).", "FR-001"),
        ("app_user", "password_hash", "VARCHAR", "NOT NULL", "Hash bcrypt/argon2; nunca texto plano.", "FR-001"),
        ("app_user", "full_name", "VARCHAR(255)", "NULL", "Nombre para mostrar.", "—"),
        ("app_user", "is_active", "BOOLEAN", "Default true", "Permite desactivar sin borrar.", "FR-002"),
        ("app_user", "failed_login_count", "INT", "Default 0", "Intentos fallidos para bloqueo.", "FR-001"),
        ("app_user", "locked_until", "TIMESTAMPTZ", "NULL", "Bloqueo temporal de cuenta.", "FR-001"),
        ("app_user", "created_at", "TIMESTAMPTZ", "Default now()", "Alta del registro.", "—"),
        ("app_user", "updated_at", "TIMESTAMPTZ", "NULL", "Última modificación.", "—"),
        # role
        ("role", "id", "UUID", "PK", "Rol del sistema (catálogo maestro de roles).", "FR-004"),
        ("role", "code", "VARCHAR(32)", "UK, NOT NULL", "Código estable: admin, logistics, ar, etc.", "FR-004"),
        ("role", "name", "VARCHAR(100)", "NOT NULL", "Etiqueta humana del rol.", "FR-004"),
        # user_role
        ("user_role", "user_id", "UUID", "FK→app_user, PK compuesta", "Usuario.", "FR-002"),
        ("user_role", "role_id", "UUID", "FK→role, PK compuesta", "Rol asignado.", "FR-002"),
        # refresh_token
        ("refresh_token", "id", "UUID", "PK", "Token de refresco de sesión.", "FR-001"),
        ("refresh_token", "user_id", "UUID", "FK→app_user, NOT NULL", "Propietario del token.", "FR-001"),
        ("refresh_token", "token_hash", "VARCHAR", "NOT NULL", "Hash del token; no guardar JWT en claro.", "FR-001"),
        ("refresh_token", "expires_at", "TIMESTAMPTZ", "NOT NULL", "Expiración.", "FR-001"),
        ("refresh_token", "revoked_at", "TIMESTAMPTZ", "NULL", "Revocación (logout).", "FR-001"),
        ("refresh_token", "created_at", "TIMESTAMPTZ", "Default now()", "Emisión.", "—"),
        # password_reset_token
        ("password_reset_token", "id", "UUID", "PK", "Token único restablecimiento.", "FR-001"),
        ("password_reset_token", "user_id", "UUID", "FK→app_user", "Usuario destino.", "FR-001"),
        ("password_reset_token", "token_hash", "VARCHAR", "NOT NULL", "Hash del token enviado por email.", "FR-001"),
        ("password_reset_token", "expires_at", "TIMESTAMPTZ", "NOT NULL", "Ventana de validez.", "FR-001"),
        ("password_reset_token", "used_at", "TIMESTAMPTZ", "NULL", "Consumo único.", "FR-001"),
        ("password_reset_token", "created_at", "TIMESTAMPTZ", "Default now()", "Creación.", "—"),
        # audit_log
        ("audit_log", "id", "BIGSERIAL", "PK", "Secuencia append-only de auditoría.", "FR-003"),
        ("audit_log", "tenant_id", "UUID", "FK→tenant, NULL", "Instancia (si aplica).", "FR-003"),
        ("audit_log", "user_id", "UUID", "FK→app_user, NULL", "Actor.", "FR-003"),
        ("audit_log", "entity_type", "VARCHAR(80)", "NOT NULL", "Tipo entidad: product, shipment, import_batch…", "FR-003"),
        ("audit_log", "entity_id", "UUID", "NULL", "PK UUID de la entidad.", "FR-003"),
        ("audit_log", "entity_id_bigint", "BIGINT", "NULL", "Alternativa si la entidad usa PK numérica.", "FR-003"),
        ("audit_log", "action", "VARCHAR(64)", "NOT NULL", "create, update, approve, etc.", "FR-003"),
        ("audit_log", "payload_json", "JSONB", "NULL", "Diff o snapshot JSON.", "FR-003"),
        ("audit_log", "created_at", "TIMESTAMPTZ", "NOT NULL, default now()", "Momento del evento.", "FR-003"),
        # client
        ("client", "id", "UUID", "PK", "Cliente constructor / empresa.", "FR-020"),
        ("client", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Multi-instancia.", "FR-020"),
        ("client", "nit", "VARCHAR(32)", "NOT NULL, UK(tenant,nit)", "NIT; formato único acordado.", "INDEX.md / FR-020"),
        ("client", "legal_name", "VARCHAR(255)", "NOT NULL", "Razón social.", "FR-020"),
        ("client", "remission_prefix", "VARCHAR(16)", "NULL", "Siglas para código remisión visible (FR-041).", "FR-041"),
        ("client", "is_active", "BOOLEAN", "Default true", "Cliente operativo.", "—"),
        ("client", "created_at", "TIMESTAMPTZ", "Default now()", "Alta.", "—"),
        # client_branch
        ("client_branch", "id", "UUID", "PK", "Sucursal del cliente (SUC).", "FR-020"),
        ("client_branch", "client_id", "UUID", "FK→client, NOT NULL", "Cliente padre.", "FR-020"),
        ("client_branch", "suc_code", "VARCHAR(16)", "NOT NULL, UK(client,suc)", "Código sucursal tipo reporte INACAR.", "INDEX.md"),
        ("client_branch", "name", "VARCHAR(255)", "NULL", "Nombre sucursal.", "—"),
        # site
        ("site", "id", "UUID", "PK", "Obra / proyecto / destino.", "FR-021"),
        ("site", "client_branch_id", "UUID", "FK→client_branch, NOT NULL", "Sucursal a la que pertenece la obra.", "FR-021"),
        ("site", "name", "VARCHAR(255)", "NOT NULL", "Nombre obra.", "FR-021"),
        ("site", "address", "TEXT", "NULL", "Dirección.", "FR-021"),
        ("site", "contact_name", "VARCHAR(255)", "NULL", "Contacto en sitio.", "FR-021"),
        ("site", "contact_phone", "VARCHAR(64)", "NULL", "Teléfono contacto.", "FR-021"),
        ("site", "created_at", "TIMESTAMPTZ", "Default now()", "Creación.", "—"),
        # warehouse
        ("warehouse", "id", "UUID", "PK", "Bodega física o lógica.", "FR-012"),
        ("warehouse", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-012"),
        ("warehouse", "code", "VARCHAR(32)", "NOT NULL, UK(tenant,code)", "Código bodega (ej. Engativa, 50 clientes).", "INDEX.md"),
        ("warehouse", "name", "VARCHAR(255)", "NOT NULL", "Nombre legible.", "FR-012"),
        # product
        ("product", "id", "UUID", "PK", "Producto catálogo Alforequipos.", "FR-010"),
        ("product", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-010"),
        ("product", "sku_code", "VARCHAR(64)", "NOT NULL, UK(tenant,sku)", "Código SKU alineado a Siigo/Excel.", "FR-010"),
        ("product", "description", "TEXT", "NOT NULL", "Descripción del ítem.", "FR-010"),
        ("product", "category", "VARCHAR(128)", "NULL", "Clasificación.", "FR-010"),
        ("product", "rental_enabled", "BOOLEAN", "Default false", "Permite alquiler.", "FR-010"),
        ("product", "sale_enabled", "BOOLEAN", "Default false", "Permite venta directa.", "FR-010"),
        ("product", "rental_price_per_day", "DECIMAL(18,4)", "NULL", "Precio alquiler por DÍA (no hora).", "FR-011, call2"),
        ("product", "sale_price", "DECIMAL(18,4)", "NULL", "Precio venta si aplica.", "FR-011"),
        ("product", "usage_notes", "TEXT", "NULL", "Uso / observaciones FR-010a; sin imagen MVP.", "FR-010a"),
        ("product", "is_active", "BOOLEAN", "Default true", "Ítem activo en operación.", "—"),
        ("product", "created_at", "TIMESTAMPTZ", "Default now()", "Alta.", "—"),
        ("product", "updated_at", "TIMESTAMPTZ", "NULL", "Última modificación.", "—"),
        # product_warehouse_stock
        ("product_warehouse_stock", "product_id", "UUID", "FK→product, PK comp.", "Producto.", "FR-012"),
        ("product_warehouse_stock", "warehouse_id", "UUID", "FK→warehouse, PK comp.", "Bodega.", "FR-012"),
        ("product_warehouse_stock", "quantity", "DECIMAL(18,4)", "Default 0", "Existencia por bodega.", "FR-012"),
        # import_batch
        ("import_batch", "id", "UUID", "PK", "Lote de importación Excel.", "FR-013"),
        ("import_batch", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-013"),
        ("import_batch", "created_by", "UUID", "FK→app_user, NULL", "Usuario que subió el archivo.", "FR-013"),
        ("import_batch", "status", "VARCHAR(32)", "NOT NULL, default draft", "draft|pending_approval|approved|rejected.", "FR-014"),
        ("import_batch", "file_name", "VARCHAR(512)", "NULL", "Nombre archivo origen.", "FR-013"),
        ("import_batch", "total_rows", "INT", "Default 0", "Total filas procesadas.", "FR-013"),
        ("import_batch", "error_rows", "INT", "Default 0", "Filas con error.", "FR-013"),
        ("import_batch", "created_at", "TIMESTAMPTZ", "Default now()", "Carga.", "—"),
        ("import_batch", "approved_at", "TIMESTAMPTZ", "NULL", "Aprobación admin.", "FR-014"),
        ("import_batch", "approved_by", "UUID", "FK→app_user, NULL", "Usuario que aprobó.", "FR-014"),
        # import_batch_row
        ("import_batch_row", "id", "BIGSERIAL", "PK", "Detalle por fila del Excel.", "FR-013"),
        ("import_batch_row", "batch_id", "UUID", "FK→import_batch, NOT NULL", "Lote padre.", "FR-013"),
        ("import_batch_row", "row_number", "INT", "NOT NULL", "Número de fila en archivo.", "FR-013"),
        ("import_batch_row", "raw_json", "JSONB", "NOT NULL", "Payload parseado.", "FR-013"),
        ("import_batch_row", "status", "VARCHAR(32)", "NULL", "ok|error por fila.", "FR-013"),
        ("import_batch_row", "error_message", "TEXT", "NULL", "Mensaje validación.", "FR-013"),
        # credit_application
        ("credit_application", "id", "UUID", "PK", "Solicitud de crédito.", "FR-022"),
        ("credit_application", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-022"),
        ("credit_application", "client_id", "UUID", "FK→client, NOT NULL", "Cliente evaluado.", "FR-022"),
        ("credit_application", "status", "VARCHAR(32)", "NOT NULL, default in_review", "Estados crédito.", "FR-022"),
        ("credit_application", "notes", "TEXT", "NULL", "Notas analista.", "FR-022"),
        ("credit_application", "blocks_shipment", "BOOLEAN", "Default false", "Bloquea remisión si política aplica.", "FR-022"),
        ("credit_application", "created_at", "TIMESTAMPTZ", "Default now()", "Ingreso solicitud.", "—"),
        ("credit_application", "updated_at", "TIMESTAMPTZ", "NULL", "Última actualización.", "—"),
        # credit_application_file
        ("credit_application_file", "id", "UUID", "PK", "Adjunto documental.", "FR-022"),
        ("credit_application_file", "application_id", "UUID", "FK→credit_application, NOT NULL", "Expediente.", "FR-022"),
        ("credit_application_file", "file_url", "TEXT", "NOT NULL", "URL almacenamiento objeto.", "FR-022"),
        ("credit_application_file", "file_name", "VARCHAR(512)", "NULL", "Nombre original.", "FR-022"),
        ("credit_application_file", "uploaded_at", "TIMESTAMPTZ", "Default now()", "Carga.", "—"),
        # ticket
        ("ticket", "id", "UUID", "PK", "Solicitud / cotización.", "FR-030"),
        ("ticket", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-030"),
        ("ticket", "client_id", "UUID", "FK→client, NOT NULL", "Cliente solicitante.", "FR-030"),
        ("ticket", "site_id", "UUID", "FK→site, NULL", "Obra destino.", "FR-030"),
        ("ticket", "created_by_client_user_id", "UUID", "FK→app_user, NULL", "Usuario portal cliente si aplica.", "FR-030"),
        ("ticket", "status", "VARCHAR(32)", "NOT NULL, default received", "Máquina de estados cotización.", "FR-031"),
        ("ticket", "requested_on", "DATE", "NULL", "Fecha requerida en obra.", "FR-030"),
        ("ticket", "first_response_at", "TIMESTAMPTZ", "NULL", "Marca SLA primera respuesta.", "FR-033"),
        ("ticket", "quotation_sent_at", "TIMESTAMPTZ", "NULL", "Envío cotización.", "FR-031"),
        ("ticket", "created_at", "TIMESTAMPTZ", "Default now()", "Creación ticket.", "—"),
        ("ticket", "updated_at", "TIMESTAMPTZ", "NULL", "Último cambio.", "—"),
        # ticket_line
        ("ticket_line", "id", "UUID", "PK", "Línea de solicitud.", "FR-030"),
        ("ticket_line", "ticket_id", "UUID", "FK→ticket, NOT NULL", "Ticket padre.", "FR-030"),
        ("ticket_line", "product_id", "UUID", "FK→product, NOT NULL", "Ítem.", "FR-030"),
        ("ticket_line", "quantity", "DECIMAL(18,4)", "NOT NULL", "Cantidad solicitada.", "FR-030"),
        # ticket_attachment
        ("ticket_attachment", "id", "UUID", "PK", "Adjunto en ticket.", "FR-032"),
        ("ticket_attachment", "ticket_id", "UUID", "FK→ticket, NOT NULL", "Ticket.", "FR-032"),
        ("ticket_attachment", "file_url", "TEXT", "NOT NULL", "URL archivo.", "FR-032"),
        ("ticket_attachment", "file_name", "VARCHAR(512)", "NULL", "Nombre archivo.", "FR-032"),
        ("ticket_attachment", "uploaded_at", "TIMESTAMPTZ", "Default now()", "Subida.", "—"),
        # shipment
        ("shipment", "id", "UUID", "PK", "Remisión de envío (alquiler).", "FR-040"),
        ("shipment", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-040"),
        ("shipment", "display_code", "VARCHAR(128)", "UK, NOT NULL", "Código visible: siglas+timestamp FR-041.", "FR-041"),
        ("shipment", "client_id", "UUID", "FK→client, NOT NULL", "Cliente.", "FR-040"),
        ("shipment", "site_id", "UUID", "FK→site, NULL", "Obra.", "FR-040"),
        ("shipment", "ticket_id", "UUID", "FK→ticket, NULL", "Origen solicitud si aplica.", "FR-040"),
        ("shipment", "kind", "VARCHAR(32)", "Default rental_dispatch", "Tipo remisión.", "FR-040"),
        ("shipment", "occurred_at", "TIMESTAMPTZ", "NOT NULL", "Momento base para código / trazabilidad.", "FR-041"),
        ("shipment", "created_by", "UUID", "FK→app_user, NULL", "Usuario logística.", "FR-040"),
        ("shipment", "created_at", "TIMESTAMPTZ", "Default now()", "Registro sistema.", "—"),
        # shipment_line
        ("shipment_line", "id", "UUID", "PK", "Línea remisión.", "FR-042"),
        ("shipment_line", "shipment_id", "UUID", "FK→shipment, NOT NULL", "Remisión.", "FR-042"),
        ("shipment_line", "product_id", "UUID", "FK→product, NOT NULL", "Producto enviado.", "FR-042"),
        ("shipment_line", "quantity_shipped", "DECIMAL(18,4)", "NOT NULL", "Cantidad enviada.", "FR-042"),
        # direct_sale
        ("direct_sale", "id", "UUID", "PK", "Venta directa (no alquiler).", "FR-043"),
        ("direct_sale", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-043"),
        ("direct_sale", "client_id", "UUID", "FK→client, NOT NULL", "Comprador.", "FR-043"),
        ("direct_sale", "site_id", "UUID", "FK→site, NULL", "Obra opcional.", "FR-043"),
        ("direct_sale", "internal_reference", "VARCHAR(64)", "NULL", "Referencia interna orden.", "FR-043"),
        ("direct_sale", "created_by", "UUID", "FK→app_user, NULL", "Usuario que registra.", "FR-043"),
        ("direct_sale", "created_at", "TIMESTAMPTZ", "Default now()", "Fecha registro.", "—"),
        # direct_sale_line
        ("direct_sale_line", "id", "UUID", "PK", "Línea venta directa.", "FR-043"),
        ("direct_sale_line", "direct_sale_id", "UUID", "FK→direct_sale, NOT NULL", "Cabeza.", "FR-043"),
        ("direct_sale_line", "product_id", "UUID", "FK→product, NOT NULL", "Producto.", "FR-043"),
        ("direct_sale_line", "quantity", "DECIMAL(18,4)", "NOT NULL", "Cantidad.", "FR-043"),
        ("direct_sale_line", "unit_price", "DECIMAL(18,4)", "NOT NULL", "Precio unitario pactado.", "FR-043"),
        # billing_draft
        ("billing_draft", "id", "UUID", "PK", "Borrador / proforma insumo (no CPE).", "FR-070"),
        ("billing_draft", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-070"),
        ("billing_draft", "client_id", "UUID", "FK→client, NOT NULL", "Cliente a facturar en Siigo.", "FR-070"),
        ("billing_draft", "period_start", "DATE", "NOT NULL", "Inicio periodo liquidación.", "FR-070"),
        ("billing_draft", "period_end", "DATE", "NOT NULL", "Fin periodo.", "FR-070"),
        ("billing_draft", "status", "VARCHAR(32)", "NULL", "draft|client_approved|exported.", "FR-070"),
        ("billing_draft", "client_approved_at", "TIMESTAMPTZ", "NULL", "Aprobación cliente.", "FR-071"),
        ("billing_draft", "created_at", "TIMESTAMPTZ", "Default now()", "Creación borrador.", "—"),
        # billing_draft_line
        ("billing_draft_line", "id", "UUID", "PK", "Línea borrador.", "FR-070"),
        ("billing_draft_line", "draft_id", "UUID", "FK→billing_draft, NOT NULL", "Borrador.", "FR-070"),
        ("billing_draft_line", "product_id", "UUID", "FK→product, NULL", "Producto si aplica.", "FR-070"),
        ("billing_draft_line", "description", "TEXT", "NULL", "Texto libre línea.", "FR-070"),
        ("billing_draft_line", "quantity", "DECIMAL(18,4)", "NULL", "Cantidad.", "FR-070"),
        ("billing_draft_line", "unit_price", "DECIMAL(18,4)", "NULL", "Precio unitario.", "FR-070"),
        ("billing_draft_line", "line_total", "DECIMAL(18,4)", "NULL", "Total línea.", "FR-070"),
        # siigo_invoice_mirror
        ("siigo_invoice_mirror", "id", "UUID", "PK", "Espejo factura Siigo (no emite CPE).", "FR-080"),
        ("siigo_invoice_mirror", "tenant_id", "UUID", "FK→tenant, NOT NULL", "Instancia.", "FR-080"),
        ("siigo_invoice_mirror", "client_id", "UUID", "FK→client, NOT NULL", "Deudor.", "FR-080"),
        ("siigo_invoice_mirror", "siigo_document_number", "VARCHAR(64)", "UK(tenant, doc)", "Número documento Siigo.", "FR-080"),
        ("siigo_invoice_mirror", "issue_date", "DATE", "NULL", "Fecha emisión FE.", "FR-080"),
        ("siigo_invoice_mirror", "due_date", "DATE", "NULL", "Vencimiento cobro.", "FR-080"),
        ("siigo_invoice_mirror", "total_amount", "DECIMAL(18,4)", "NOT NULL", "Valor total factura.", "FR-080"),
        ("siigo_invoice_mirror", "balance_amount", "DECIMAL(18,4)", "NOT NULL", "Saldo pendiente.", "FR-080"),
        ("siigo_invoice_mirror", "status", "VARCHAR(32)", "NOT NULL", "pending|partial|paid|overdue|…", "FR-082"),
        ("siigo_invoice_mirror", "raw_payload_json", "JSONB", "NULL", "Respuesta cruda sync/import.", "FR-080"),
        ("siigo_invoice_mirror", "last_synced_at", "TIMESTAMPTZ", "NULL", "Última sincronización.", "FR-080"),
        # payment_entry
        ("payment_entry", "id", "UUID", "PK", "Registro de pago / abono.", "FR-081"),
        ("payment_entry", "invoice_id", "UUID", "FK→siigo_invoice_mirror, NOT NULL", "Factura afectada.", "FR-081"),
        ("payment_entry", "channel", "VARCHAR(64)", "NULL", "transfer|cash|portal…", "FR-081"),
        ("payment_entry", "reference", "VARCHAR(128)", "NULL", "Referencia banco/comprobante.", "FR-081"),
        ("payment_entry", "amount", "DECIMAL(18,4)", "NOT NULL", "Monto aplicado.", "FR-081"),
        ("payment_entry", "paid_at", "TIMESTAMPTZ", "NULL", "Fecha pago efectivo.", "FR-081"),
        ("payment_entry", "notes", "TEXT", "NULL", "Observaciones.", "FR-081"),
        ("payment_entry", "attachment_url", "TEXT", "NULL", "Soporte cargado.", "FR-081"),
        ("payment_entry", "created_at", "TIMESTAMPTZ", "Default now()", "Registro en sistema.", "—"),
        # ar_reminder_log
        ("ar_reminder_log", "id", "BIGSERIAL", "PK", "Historial envío recordatorios cobro.", "FR-082"),
        ("ar_reminder_log", "invoice_id", "UUID", "FK→siigo_invoice_mirror, NOT NULL", "Factura.", "FR-082"),
        ("ar_reminder_log", "sent_at", "TIMESTAMPTZ", "NOT NULL", "Momento envío.", "FR-082"),
        ("ar_reminder_log", "channel", "VARCHAR(32)", "NULL", "email, etc.", "FR-082"),
        ("ar_reminder_log", "template_version", "VARCHAR(32)", "NULL", "Versión plantilla.", "FR-082"),
        ("ar_reminder_log", "status", "VARCHAR(32)", "NULL", "delivered|failed…", "FR-082"),
    ]

    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
    autosize(ws1)

    # --- Hoja 2: Índices en modelo DBML ---
    ws2 = wb.create_sheet("Indices_Modelo_DBML")
    add_sheet_headers(
        ws2,
        ["Tabla", "Indice", "Tipo", "Columnas", "Origen"],
    )
    idx_rows = [
        ("app_user", "uq_tenant_email", "UNIQUE", "tenant_id, email", "database-model.dbml"),
        ("audit_log", "idx_entity_type_created", "INDEX", "entity_type, created_at", "database-model.dbml"),
        ("audit_log", "idx_user_created", "INDEX", "user_id, created_at", "database-model.dbml"),
        ("client", "uq_tenant_nit", "UNIQUE", "tenant_id, nit", "database-model.dbml"),
        ("client_branch", "uq_client_suc", "UNIQUE", "client_id, suc_code", "database-model.dbml"),
        ("warehouse", "uq_tenant_code", "UNIQUE", "tenant_id, code", "database-model.dbml"),
        ("product", "uq_tenant_sku", "UNIQUE", "tenant_id, sku_code", "database-model.dbml"),
        ("product_warehouse_stock", "pk_product_wh", "PRIMARY", "product_id, warehouse_id", "database-model.dbml"),
        ("user_role", "pk_user_role", "PRIMARY", "user_id, role_id", "database-model.dbml"),
        ("shipment", "uq_display_code", "UNIQUE", "display_code", "database-model.dbml"),
        ("siigo_invoice_mirror", "uq_tenant_siigo_doc", "UNIQUE", "tenant_id, siigo_document_number", "database-model.dbml"),
    ]
    for r, row in enumerate(idx_rows, start=2):
        for c, val in enumerate(row, start=1):
            ws2.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws2)

    # --- Hoja 3: Índices sugeridos (adicionales) ---
    ws3 = wb.create_sheet("Indices_Sugeridos")
    add_sheet_headers(
        ws3,
        [
            "Tabla",
            "Indice_sugerido",
            "Columnas",
            "Motivo",
            "Prioridad",
            "Notas_implementacion",
        ],
    )
    sug = [
        (
            "refresh_token",
            "idx_refresh_user_expires",
            "user_id, expires_at",
            "Listar tokens válidos por usuario y limpieza de sesión.",
            "Alta",
            "Partial index WHERE revoked_at IS NULL opcional.",
        ),
        (
            "password_reset_token",
            "idx_reset_user_expires",
            "user_id, expires_at",
            "Invalidar tokens antiguos al solicitar nuevo reset.",
            "Media",
            "—",
        ),
        (
            "import_batch",
            "idx_batch_tenant_status_created",
            "tenant_id, status, created_at DESC",
            "Bandeja de lotes pendientes por tenant.",
            "Alta",
            "—",
        ),
        (
            "import_batch_row",
            "idx_batch_row_batch",
            "batch_id, row_number",
            "Cargar errores por lote ordenado.",
            "Alta",
            "FK batch_id ya implica index en PG; compuesto acelera orden.",
        ),
        (
            "ticket",
            "idx_ticket_tenant_status_updated",
            "tenant_id, status, updated_at DESC",
            "Listados operativos y tableros.",
            "Alta",
            "—",
        ),
        (
            "ticket",
            "idx_ticket_client_created",
            "client_id, created_at DESC",
            "Historial por cliente.",
            "Media",
            "—",
        ),
        (
            "shipment",
            "idx_shipment_tenant_client_date",
            "tenant_id, client_id, occurred_at DESC",
            "Reportes movimiento y kardex por cliente/periodo.",
            "Alta",
            "PDR-04 / INDEX.md.",
        ),
        (
            "shipment_line",
            "idx_shipment_line_product",
            "product_id",
            "Agregaciones por producto (existencias en obra vía lógica).",
            "Media",
            "Evaluar MV para saldos.",
        ),
        (
            "direct_sale",
            "idx_ds_tenant_created",
            "tenant_id, created_at DESC",
            "Listado ventas recientes.",
            "Media",
            "—",
        ),
        (
            "siigo_invoice_mirror",
            "idx_ar_tenant_status_due",
            "tenant_id, status, due_date",
            "Cartera: mora y recordatorios.",
            "Alta",
            "FR-082.",
        ),
        (
            "siigo_invoice_mirror",
            "idx_ar_client_balance",
            "client_id, balance_amount",
            "Consulta saldo por cliente.",
            "Media",
            "WHERE balance_amount > 0 parcial.",
        ),
        (
            "payment_entry",
            "idx_payment_invoice_created",
            "invoice_id, created_at",
            "Historial pagos por factura.",
            "Media",
            "—",
        ),
        (
            "credit_application",
            "idx_credit_client_status",
            "client_id, status",
            "Validar bloqueo remisión antes de insert shipment.",
            "Alta",
            "Trigger o check en servicio.",
        ),
        (
            "billing_draft",
            "idx_billing_tenant_period",
            "tenant_id, period_start, period_end",
            "Evitar solapamiento de borradores por cliente (regla app).",
            "Media",
            "Unique parcial si negocio lo exige.",
        ),
    ]
    for r, row in enumerate(sug, start=2):
        for c, val in enumerate(row, start=1):
            ws3.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws3)

    # --- Hoja 4: Rutinas sugeridas (otras horas) ---
    ws4 = wb.create_sheet("Rutinas_Sugeridas")
    add_sheet_headers(
        ws4,
        [
            "Nombre",
            "Tipo",
            "Objetivo",
            "Dependencias",
            "Estimacion_horas",
            "Notas",
        ],
    )
    routines = [
        (
            "fn_generate_shipment_display_code",
            "FUNCTION (SQL/PLpgSQL)",
            "Generar display_code único = upper(client.remission_prefix) || '_' || to_char(occurred_at AT TIME ZONE tenant_tz, 'YYYYMMDDHH24MISS').",
            "client.remission_prefix, shipment.occurred_at, tenant.timezone",
            "4–6",
            "Validar unicidad; colisión → retry con sufijo.",
        ),
        (
            "trg_audit_shipment",
            "TRIGGER",
            "Insertar en audit_log tras INSERT/UPDATE en shipment y shipment_line.",
            "audit_log, shipment*",
            "3–4",
            "FR-003; no registrar lecturas.",
        ),
        (
            "trg_audit_product_import",
            "TRIGGER",
            "Auditar cambio de estado import_batch a approved/rejected.",
            "import_batch, audit_log",
            "2–3",
            "—",
        ),
        (
            "job_sync_siigo_invoices",
            "JOB (Celery/Redis)",
            "Sincronizar/importar facturas desde API o archivo hacia siigo_invoice_mirror.",
            "Siigo, credenciales, cola",
            "16–24",
            "Spike integración aparte; ver pdr/QUESTIONS.md.",
        ),
        (
            "job_ar_reminders",
            "JOB programado",
            "Enviar recordatorios email facturas vencidas / mora; escribir ar_reminder_log.",
            "siigo_invoice_mirror, SMTP",
            "8–12",
            "FR-082; plantillas legales.",
        ),
        (
            "mv_refresh_kardex_client",
            "MATERIALIZED VIEW + REFRESH",
            "Precomputar saldos por cliente/producto según shipment_line y direct_sale_line (MVP sin devolución).",
            "shipment*, product*, client*",
            "12–20",
            "CONCURRENTLY en ventana mantenimiento; PDR-04.",
        ),
        (
            "vw_movement_detail_export",
            "VIEW",
            "Vista compatible columnas INACAR (NIT, SUC, remisión, devolución=0, saldo).",
            "Varias tablas",
            "6–10",
            "Solo lectura para export API.",
        ),
        (
            "sp_apply_payment_to_invoice",
            "PROCEDURE",
            "Actualizar balance_amount y status en siigo_invoice_mirror al registrar payment_entry.",
            "payment_entry, siigo_invoice_mirror",
            "4–6",
            "Transacción única; evitar race.",
        ),
        (
            "fn_soft_delete_guard",
            "TRIGGER",
            "Impedir DELETE físico en tablas críticas (opcional) o forzar soft-delete.",
            "product, shipment",
            "3–5",
            "Alineado NFR integridad.",
        ),
        (
            "maintenance_cleanup_tokens",
            "CRON/SQL",
            "Purgar refresh_token y password_reset_token expirados.",
            "refresh_token, password_reset_token",
            "2",
            "Mantenimiento semanal.",
        ),
    ]
    for r, row in enumerate(routines, start=2):
        for c, val in enumerate(row, start=1):
            ws4.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws4)

    # --- Hoja 5: Portada / metadatos ---
    ws0 = wb.create_sheet("Leeme", 0)
    ws0["A1"] = "Diccionario de datos — G&H / gyhobras"
    ws0["A2"] = "Fuente: database-model.dbml | PostgreSQL relacional"
    ws0["A3"] = "Generado por: scripts/generate_data_dictionary_excel.py"
    ws0["A5"] = "Hojas:"
    ws0["A6"] = "1) Diccionario_Campos — columnas y descripciones"
    ws0["A7"] = "2) Indices_Modelo_DBML — índices ya definidos en el .dbml"
    ws0["A8"] = "3) Indices_Sugeridos — índices adicionales para rendimiento"
    ws0["A9"] = "4) Rutinas_Sugeridas — funciones, triggers, jobs (estimación en horas aparte)"
    ws0["A11"] = "Nota: Las estimaciones de rutinas son orientativas para planificación (otras horas)."
    for row in range(1, 12):
        ws0.cell(row=row, column=1).alignment = WRAP
    ws0.column_dimensions["A"].width = 85

    wb.save(OUT)
    print(f"Escrito: {OUT}")


if __name__ == "__main__":
    main()
